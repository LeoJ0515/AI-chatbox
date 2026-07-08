import os
import json
import base64
import io
import uuid
import hashlib
import secrets
import datetime
from fastapi import FastAPI, HTTPException, Depends, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from groq import Groq
from supabase import create_client, Client
from dotenv import load_dotenv
from typing import List, Dict, Optional

# Document parsing imports
import PyPDF2
from docx import Document
from openpyxl import load_workbook

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")           # service_role key
GROQ_API_KEY = os.getenv("GROQ_API_KEY")

if not all([SUPABASE_URL, SUPABASE_KEY, GROQ_API_KEY]):
    raise RuntimeError("Missing required environment variables")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
groq_client = Groq(api_key=GROQ_API_KEY)

GROQ_MODEL = "llama-3.3-70b-versatile"
SYSTEM_PROMPT = (
    "You are a helpful, professional AI assistant. "
    "Answer questions accurately and in detail. "
    "If you receive file content, use it to answer the user's query."
)
MAX_HISTORY_CHARS = 50000

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# -------------------------------------------------------------------
# Pydantic models
# -------------------------------------------------------------------
class FileAttachment(BaseModel):
    name: str
    type: str
    data: str  # base64

class ChatRequest(BaseModel):
    session_id: str
    message: str
    files: Optional[List[FileAttachment]] = []

class ChatResponse(BaseModel):
    reply: str

class AuthSignup(BaseModel):
    username: str
    email: str
    password: str

class AuthLogin(BaseModel):
    email: str
    password: str

# -------------------------------------------------------------------
# Custom User Authentication
# -------------------------------------------------------------------
def hash_password(password: str) -> str:
    """Hash a password with SHA-256 + salt."""
    salt = secrets.token_hex(16)
    hashed = hashlib.sha256((password + salt).encode()).hexdigest()
    return f"{salt}${hashed}"

def verify_password(password: str, stored_hash: str) -> bool:
    """Verify a password against stored hash."""
    salt, hashed = stored_hash.split("$")
    return hashlib.sha256((password + salt).encode()).hexdigest() == hashed

def create_token(user_id: str) -> str:
    # Remove any existing tokens for this user
    supabase.table("user_tokens") \
            .delete() \
            .eq("user_id", user_id) \
            .execute()

    # Create a fresh token
    token = secrets.token_hex(32)
    expires = datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(days=7)
    supabase.table("user_tokens").insert({
        "user_id": user_id,
        "token": token,
        "expires_at": expires.isoformat()
    }).execute()
    return token

def verify_token(token: str) -> Optional[str]:
    """Verify a token and return user_id if valid."""
    res = supabase.table("user_tokens") \
                  .select("*") \
                  .eq("token", token) \
                  .execute()
    if not res.data:
        return None
    record = res.data[0]
    expires_at = datetime.datetime.fromisoformat(record["expires_at"])
    if expires_at < datetime.datetime.now(datetime.timezone.utc):
        supabase.table("user_tokens").delete().eq("id", record["id"]).execute()
        return None
    return record["user_id"]

async def get_current_user(authorization: str = Header(None)):
    """Verify custom JWT and return user ID."""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing or invalid token")
    token = authorization.split(" ")[1]
    user_id = verify_token(token)
    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid token")
    return user_id

# -------------------------------------------------------------------
# Database helpers
# -------------------------------------------------------------------
def load_messages(user_id: str, session_id: str) -> List[Dict[str, str]]:
    res = supabase.table("conversations") \
                  .select("messages") \
                  .eq("user_id", user_id) \
                  .eq("session_id", session_id) \
                  .execute()
    return res.data[0]["messages"] if res.data else []

def save_messages(user_id: str, session_id: str, messages: List[Dict[str, str]]):
    supabase.table("conversations") \
            .upsert({
                "user_id": user_id,
                "session_id": session_id,
                "messages": messages,
                "updated_at": "now()"
            }, on_conflict="user_id,session_id") \
            .execute()

def trim_history(messages: List[Dict]) -> List[Dict]:
    total = sum(len(json.dumps(m)) for m in messages)
    while total > MAX_HISTORY_CHARS and len(messages) > 2:
        removed = messages.pop(0)
        if messages and messages[0]["role"] == "assistant":
            removed = messages.pop(0)
        total = sum(len(json.dumps(m)) for m in messages)
    return messages

def call_groq(messages: List[Dict]) -> str:
    api_messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    for msg in messages:
        api_messages.append({"role": msg["role"], "content": msg["content"]})
    response = groq_client.chat.completions.create(
        model=GROQ_MODEL,
        messages=api_messages,
        temperature=0.7,
        max_tokens=4096,
    )
    return response.choices[0].message.content

def extract_file_content(file: FileAttachment) -> str:
    ext = file.name.split('.')[-1].lower() if '.' in file.name else ''
    try:
        file_bytes = base64.b64decode(file.data)
    except Exception:
        return f"[Error decoding file {file.name}]"

    text_extensions = {'txt', 'csv', 'md', 'json', 'xml', 'yaml', 'yml', 'py', 'js',
                       'ts', 'html', 'css', 'sh', 'bat', 'log', 'ini', 'cfg', 'toml'}
    text_mime_prefixes = ["text/", "application/json", "application/javascript",
                          "application/xml", "application/x-python", "application/x-sh"]
    is_text = any(file.type.startswith(p) for p in text_mime_prefixes) or ext in text_extensions

    if is_text:
        try:
            text = file_bytes.decode('utf-8', errors='replace')
            return f"--- Content of {file.name} ---\n{text}\n--- End of file ---"
        except Exception as e:
            return f"[Error reading text file {file.name}: {str(e)}]"

    if ext == 'pdf' or file.type == 'application/pdf':
        try:
            reader = PyPDF2.PdfReader(io.BytesIO(file_bytes))
            text = ""
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            if not text.strip():
                return f"[PDF {file.name} contains no extractable text]"
            return f"--- Content of {file.name} (PDF) ---\n{text}\n--- End of file ---"
        except Exception as e:
            return f"[Error reading PDF {file.name}: {str(e)}]"

    if ext == 'docx' or file.type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
        try:
            doc = Document(io.BytesIO(file_bytes))
            text = "\n".join([para.text for para in doc.paragraphs])
            if not text.strip():
                text = "[Document appears empty]"
            return f"--- Content of {file.name} (Word) ---\n{text}\n--- End of file ---"
        except Exception as e:
            return f"[Error reading Word file {file.name}: {str(e)}]"

    if ext == 'xlsx' or file.type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            text = ""
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in ws.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    text += row_text + "\n"
                text += "\n"
            wb.close()
            return f"--- Content of {file.name} (Excel) ---\n{text}\n--- End of file ---"
        except Exception as e:
            return f"[Error reading Excel file {file.name}: {str(e)}]"

    if file.type.startswith("image/"):
        return f"[Image attached: {file.name}]"
    return f"[File attached: {file.name} (unsupported type)]"

# -------------------------------------------------------------------
# AUTH ENDPOINTS (Custom - No Supabase Auth)
# -------------------------------------------------------------------

# ---- Signup ----
@app.post("/auth/signup")
async def signup(data: AuthSignup):
    # Validate
    if not data.username or len(data.username) < 3:
        raise HTTPException(status_code=400, detail="Username must be at least 3 characters")
    if not data.email or "@" not in data.email:
        raise HTTPException(status_code=400, detail="Invalid email")
    if not data.password or len(data.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")

    # Check if email already exists
    existing = supabase.table("users") \
                       .select("id") \
                       .eq("email", data.email) \
                       .execute()
    if existing.data:
        raise HTTPException(status_code=400, detail="Email already registered")

    # Check if username already exists
    existing_user = supabase.table("users") \
                            .select("id") \
                            .eq("username", data.username) \
                            .execute()
    if existing_user.data:
        raise HTTPException(status_code=400, detail="Username already taken")

    # Create user
    user_id = str(uuid.uuid4())
    hashed = hash_password(data.password)

    supabase.table("users").insert({
        "id": user_id,
        "username": data.username,
        "email": data.email,
        "password_hash": hashed,
        "created_at": datetime.datetime.now(datetime.timezone.utc).isoformat()
    }).execute()

    return {"message": "Signup successful. You can now login."}

# ---- Login ----
@app.post("/auth/login")
async def login(data: AuthLogin):
    # Find user by email
    res = supabase.table("users") \
                  .select("*") \
                  .eq("email", data.email) \
                  .execute()

    if not res.data:
        raise HTTPException(status_code=401, detail="Invalid credentials")

    user = res.data[0]

    # Verify password
    if not verify_password(data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Create token
    token = create_token(user["id"])

    return {
        "access_token": token,
        "token_type": "bearer",
        "username": user["username"]
    }

# -------------------------------------------------------------------
# PROTECTED CHAT ENDPOINTS
# -------------------------------------------------------------------
@app.post("/chat", response_model=ChatResponse)
async def chat(req: ChatRequest, user_id: str = Depends(get_current_user)):
    session_id = req.session_id.strip()
    user_message = req.message.strip()

    file_contents = []
    if req.files:
        for f in req.files:
            content = extract_file_content(f)
            if content:
                file_contents.append(content)

    effective_message = user_message
    if file_contents:
        file_text_block = "\n\n".join(file_contents)
        effective_message = f"{user_message}\n\n[Attached file content]\n{file_text_block}"

    if not effective_message:
        raise HTTPException(status_code=400, detail="Empty message")

    messages = load_messages(user_id, session_id)
    messages.append({"role": "user", "content": effective_message})
    messages = trim_history(messages)

    try:
        assistant_reply = call_groq(messages)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LLM error: {str(e)}")

    messages.append({"role": "assistant", "content": assistant_reply})
    save_messages(user_id, session_id, messages)

    return ChatResponse(reply=assistant_reply)

@app.get("/chat/{session_id}")
async def get_messages(session_id: str, user_id: str = Depends(get_current_user)):
    messages = load_messages(user_id, session_id)
    return {"messages": messages}

@app.delete("/chat/{session_id}")
async def clear_history(session_id: str, user_id: str = Depends(get_current_user)):
    supabase.table("conversations") \
            .delete() \
            .eq("user_id", user_id) \
            .eq("session_id", session_id) \
            .execute()
    return {"message": "History cleared"}

# ── Pydantic model for session response ──
class SessionInfo(BaseModel):
    id: str
    title: str
    lastActive: str

@app.get("/sessions", response_model=List[SessionInfo])
async def get_sessions(user_id: str = Depends(get_current_user)):
    res = supabase.table("user_sessions") \
                  .select("id, title, updated_at") \
                  .eq("user_id", user_id) \
                  .order("updated_at", desc=True) \
                  .execute()
    sessions = []
    for row in res.data:
        sessions.append(SessionInfo(
            id=row["id"],
            title=row["title"],
            lastActive=row["updated_at"]
        ))
    return sessions

@app.post("/sessions")
async def create_session(user_id: str = Depends(get_current_user)):
    new_id = str(uuid.uuid4())
    supabase.table("user_sessions").insert({
        "id": new_id,
        "user_id": user_id,
        "title": "New Session",
        "updated_at": datetime.datetime.now(datetime.timezone.utc).isoformat()
    }).execute()
    return {"id": new_id, "title": "New Session", "lastActive": datetime.datetime.now(datetime.timezone.utc).isoformat()}

@app.put("/sessions/{session_id}")
async def update_session_title(session_id: str, title: str, user_id: str = Depends(get_current_user)):
    supabase.table("user_sessions") \
            .update({"title": title, "updated_at": "now()"}) \
            .eq("id", session_id) \
            .eq("user_id", user_id) \
            .execute()
    return {"message": "updated"}

@app.delete("/sessions/{session_id}")
async def delete_session(session_id: str, user_id: str = Depends(get_current_user)):
    # Also delete the conversation
    supabase.table("conversations") \
            .delete() \
            .eq("user_id", user_id) \
            .eq("session_id", session_id) \
            .execute()
    supabase.table("user_sessions") \
            .delete() \
            .eq("id", session_id) \
            .eq("user_id", user_id) \
            .execute()
    return {"message": "deleted"}

# -------------------------------------------------------------------
# Frontend
# -------------------------------------------------------------------
@app.get("/")
async def root():
    return FileResponse("index.html")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)